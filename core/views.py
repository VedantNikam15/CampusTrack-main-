from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, update_session_auth_hash, authenticate
from django.contrib.auth.views import PasswordChangeView
from django.contrib.auth.decorators import login_required
from .models import Post, Certificate, Event, Marks, Notification, User, Comment
from .forms import (
    UserRegisterForm, PostForm, CertificateForm, MarksForm, CommentForm, EventForm,
    UserEditForm, StudentProfileForm, TeacherProfileForm,
)
import openpyxl
from io import BytesIO
from django.db.models import Avg, F, Q
from django.utils import timezone
from django.http import JsonResponse
from datetime import timedelta
from django.db import transaction
import re
import re
from django.contrib.auth import get_user_model
from django.db.models import Avg
from collections import defaultdict, OrderedDict
from django.core.mail import send_mail
from django.conf import settings
from django.shortcuts import HttpResponse
from .forms import EventForm
from django.contrib import messages
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.template.loader import render_to_string
from .models import News
from .forms import NewsForm


def is_approved_teacher(user):
    """Return True if the given user is an approved teacher or staff."""
    try:
        return user.is_staff or (getattr(user, 'role', None) == 'teacher' and bool(getattr(user, 'teacher_approved', False)))
    except Exception:
        return False


class CustomPasswordChangeView(PasswordChangeView):
    """Override PasswordChangeView to email the new password to the user after a successful change.

    Note: sending plaintext passwords by email is insecure; consider sending a notification without the password or a confirmation link instead.
    """
    def form_valid(self, form):
        # Save the new password (PasswordChangeView does this in form_valid)
        response = super().form_valid(form)
        # Ensure the session auth hash is updated so the user remains authenticated
        # with the new password and Django's session doesn't get invalidated.
        try:
            update_session_auth_hash(self.request, self.request.user)
        except Exception:
            pass
        try:
            new_password = form.cleaned_data.get('new_password1')
            user = self.request.user
            if user.email and new_password:
                subject = 'Your CampusTrack password was changed'
                message = f'Hello {user.get_full_name() or user.username},\n\nYour password was successfully changed.\n\nNew password: {new_password}\n\nIf you did not perform this change, please contact support immediately.'
                send_mail(subject, message, getattr(settings, 'DEFAULT_FROM_EMAIL', 'noreply@localhost'), [user.email], fail_silently=True)
        except Exception:
            # Avoid breaking the flow if email sending fails
            pass
        return response

def home(request):
    return render(request, 'home.html')

def register(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.username = form.cleaned_data['username']
            user.email = form.cleaned_data['email']
            user.role = form.cleaned_data['role']
            user.department = form.cleaned_data.get('department') or ''
            year = form.cleaned_data.get('year')
            if year:
                user.year = year
            user.save()
            # Authenticate the newly created user so `login()` knows which
            # authentication backend was used. When multiple backends are
            # configured, Django requires the backend attribute on the user.
            raw_password = form.cleaned_data.get('password1')
            auth_user = None
            try:
                if raw_password:
                    # Try authenticating with username first, then email.
                    auth_user = authenticate(request, username=user.username, password=raw_password)
                    if not auth_user and user.email:
                        auth_user = authenticate(request, username=user.email, password=raw_password)
            except Exception:
                auth_user = None

            if auth_user:
                login(request, auth_user)
            else:
                # Fallback: explicitly provide backend from settings (first one)
                from django.conf import settings as _settings
                backend = (_settings.AUTHENTICATION_BACKENDS[0] if getattr(_settings, 'AUTHENTICATION_BACKENDS', None) else None)
                if backend:
                    login(request, user, backend=backend)
                else:
                    # Last resort: login without backend (may raise in some configs)
                    login(request, user)
            return redirect('dashboard')
    else:
        form = UserRegisterForm()
    return render(request, 'register.html', {'form': form})


@login_required
def pending_teachers(request):
    """Admin view: list teacher accounts awaiting approval."""
    if not request.user.is_staff:
        return redirect('dashboard')
    pending = User.objects.filter(role='teacher', teacher_approved=False).order_by('date_joined')
    return render(request, 'admin/pending_teachers.html', {'pending': pending})


@login_required
def approve_teacher(request, pk):
    """Approve a teacher account so they gain teacher privileges.

    Only staff users may perform this action.
    """
    if not request.user.is_staff:
        return redirect('dashboard')
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.teacher_approved = True
        user.save()
        Notification.objects.create(user=user, content='Your teacher account has been approved by an administrator.')
        try:
            if getattr(settings, 'EMAIL_HOST', None) and user.email:
                send_mail('Teacher Account Approved', 'Your account has been approved as a teacher. You can now access teacher features.', settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=True)
        except Exception:
            pass
        messages.success(request, f'Approved teacher {user.get_full_name() or user.username}.')
        # If this was an AJAX request, return JSON so client JS can update the UI
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'action': 'approve', 'pk': user.pk})
        return redirect('core:pending_teachers')
    return render(request, 'admin/confirm_approve.html', {'user_obj': user})


@login_required
def reject_teacher(request, pk):
    """Reject a teacher request: demote role to student and notify the user.

    Only staff users may perform this action.
    """
    if not request.user.is_staff:
        return redirect('dashboard')
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        # demote to student and mark not approved
        user.role = 'student'
        user.teacher_approved = False
        user.save()
        Notification.objects.create(user=user, content='Your request to be a teacher was declined by an administrator.')
        try:
            if getattr(settings, 'EMAIL_HOST', None) and user.email:
                send_mail('Teacher Account Declined', 'Your request to be a teacher has been declined by an administrator.', settings.DEFAULT_FROM_EMAIL, [user.email], fail_silently=True)
        except Exception:
            pass
        messages.success(request, f'Rejected teacher request for {user.get_full_name() or user.username}.')
        # If AJAX, return JSON so client can update UI immediately
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'action': 'reject', 'pk': user.pk})
        return redirect('core:pending_teachers')
    return render(request, 'admin/confirm_reject.html', {'user_obj': user})

@login_required
def dashboard(request):
    user = request.user
    if user.role == 'student':
        posts = Post.objects.all().order_by('-created_at')
        events_college = Event.objects.filter(scope='college')
        events_dept = Event.objects.filter(scope='department', department=user.department)
        events = (events_college | events_dept).order_by('date_from')
        # Only consider students in the same department AND the same year as the current user
        dept_students = User.objects.filter(role='student', department=user.department, year=getattr(user, 'year', None))
        ranks = dept_students.annotate(avg_score=Avg(F('marks__marks_obtained') * 100.0 / F('marks__total_marks'))).order_by('-avg_score')
        ranks_list = [(s, getattr(s, 'avg_score') or 0) for s in ranks]
        position = None
        for idx, (stu, avg_score) in enumerate(ranks_list, start=1):
            if stu.pk == user.pk:
                position = idx
                break
        notifications = Notification.objects.filter(user=user).order_by('-created_at')[:10]
        return render(request, 'students/dashboard.html', {
            'posts': posts,'events': events,'position': position,'notifications': notifications,'ranks': ranks_list[:5],
        })
    elif user.role == 'teacher':
        # If the teacher account hasn't been approved yet, show a pending notice
        if not getattr(user, 'teacher_approved', False):
            return render(request, 'teachers/pending_approval.html')
        # Only show certificates that haven't been reviewed yet (verified=False and no feedback)
        pending_certs = Certificate.objects.filter(verified=False, feedback='')
        # By default only show active students; teachers can opt-in to see inactive via ?show_inactive=1
        show_inactive = request.GET.get('show_inactive') == '1'
        # Restrict teachers to only see students from their own department
        if show_inactive:
            students = User.objects.filter(role='student', department=user.department)
        else:
            students = User.objects.filter(role='student', department=user.department, is_active=True)
        notifications = Notification.objects.filter(user=user).order_by('-created_at')[:10]
        return render(request, 'teachers/dashboard.html', {
            'pending_certs': pending_certs,'students': students,'notifications': notifications,
            'show_inactive': show_inactive,
        })
    else:
        return redirect('home')

@login_required
def create_post(request):
    if request.method == 'POST':
        form = PostForm(request.POST, request.FILES)
        if form.is_valid():
            p = form.save(commit=False)
            p.author = request.user
            p.save()
            return redirect('dashboard')
    else:
        form = PostForm()
    return render(request, 'create_post.html', {'form': form})


@login_required
def edit_post(request, pk):
    p = get_object_or_404(Post, pk=pk)
    # Only author or staff can edit
    if not (request.user == p.author or request.user.is_staff):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'error': 'permission denied'}, status=403)
        return redirect('dashboard')
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        if not content:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'ok': False, 'error': 'empty content'}, status=400)
            messages.error(request, 'Content cannot be empty.')
            return redirect('dashboard')
        p.content = content
        p.save()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string('includes/post_item.html', {'post': p, 'user': request.user}, request=request)
            return JsonResponse({'ok': True, 'html': html, 'pk': p.pk})
        return redirect('dashboard')
    # GET fallback: redirect to dashboard (or show an edit page if desired)
    return redirect('dashboard')


@login_required
def delete_post(request, pk):
    p = get_object_or_404(Post, pk=pk)
    # Only author or staff can delete
    if not (request.user == p.author or request.user.is_staff):
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'error': 'permission denied'}, status=403)
        return redirect('dashboard')
    if request.method == 'POST':
        p.delete()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'deleted_pk': pk})
        return redirect('dashboard')
    return render(request, 'teachers/confirm_delete_event.html', {'event': p})

@login_required
def toggle_like(request, pk):
    post = get_object_or_404(Post, pk=pk)
    user = request.user
    if user in post.likes.all():
        post.likes.remove(user)
    else:
        post.likes.add(user)
        # Notify the post author (but not if they liked their own post)
        try:
            if post.author and post.author != user:
                preview = (post.content[:30] + '...') if post.content and len(post.content) > 30 else (post.content or '')
                Notification.objects.create(user=post.author, content=f'{user.get_full_name() or user.username} liked your post "{preview}"')
        except Exception:
            # ensure like still succeeds even if notification fails
            pass
    return redirect('dashboard')

@login_required
def add_comment(request, pk):
    post = get_object_or_404(Post, pk=pk)
    if request.method == 'POST':
        form = CommentForm(request.POST)
        if form.is_valid():
            c = form.save(commit=False)
            c.post = post
            c.author = request.user
            # Normalize whitespace for duplicate detection
            norm_content = re.sub(r"\s+", " ", (c.content or '').strip())
            # Use a DB transaction and select_for_update on the post to avoid race conditions
            try:
                with transaction.atomic():
                    # lock the post row so concurrent comment submissions serialize
                    _ = Post.objects.select_for_update().get(pk=post.pk)
                    recent_window = timezone.now() - timedelta(seconds=30)
                    dup_exists = Comment.objects.filter(post=post, author=request.user).filter(created_at__gte=recent_window).filter(content__iregex=r"^%s$" % re.escape(norm_content)).exists()
                    if dup_exists:
                        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                            return JsonResponse({'ok': True, 'html': ''})
                        return redirect('dashboard')
                    # assign normalized content before saving
                    c.content = norm_content
                    c.save()
            except Exception:
                # fallback to naive save if locking/checking fails for some reason
                c.content = norm_content
                c.save()
                
            # Notify the post author about the new comment (don't notify self)
            try:
                if post.author and post.author != request.user:
                    Notification.objects.create(user=post.author, content=f'{request.user.get_full_name() or request.user.username} commented on your post')
            except Exception:
                pass
            # If AJAX request, return rendered HTML for the single comment
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                html = render_to_string('includes/comment_item.html', {'comment': c, 'user': request.user}, request=request)
                return JsonResponse({'ok': True, 'html': html, 'id': c.id})
    return redirect('dashboard')


@login_required
def edit_comment(request, pk):
    c = get_object_or_404(Comment, pk=pk)
    # only author or staff can edit
    if not (request.user == c.author or request.user.is_staff):
        return JsonResponse({'ok': False, 'error': 'permission denied'}, status=403)
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        if not content:
            return JsonResponse({'ok': False, 'error': 'Empty content'}, status=400)
        c.content = content
        c.save()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            html = render_to_string('includes/comment_item.html', {'comment': c, 'user': request.user}, request=request)
            return JsonResponse({'ok': True, 'html': html, 'id': c.id})
    return JsonResponse({'ok': False, 'error': 'POST required'}, status=400)


@login_required
def delete_comment(request, pk):
    c = get_object_or_404(Comment, pk=pk)
    if not (request.user == c.author or request.user.is_staff):
        return JsonResponse({'ok': False, 'error': 'permission denied'}, status=403)
    if request.method == 'POST':
        cid = c.id
        c.delete()
        return JsonResponse({'ok': True, 'deleted_id': cid})
    return JsonResponse({'ok': False, 'error': 'POST required'}, status=400)

@login_required
def upload_certificate(request):
    if request.method == 'POST':
        form = CertificateForm(request.POST, request.FILES)
        if form.is_valid():
            cert = form.save(commit=False)
            cert.student = request.user
            cert.save()
            # Notify teachers that a new certificate has been uploaded for review
            teachers = User.objects.filter(role='teacher')
            content = f'Certificate uploaded by {request.user.get_full_name() or request.user.email}: "{cert.title}"'
            for t in teachers:
                Notification.objects.create(user=t, content=content)
                try:
                    if getattr(settings, 'EMAIL_HOST', None) and t.email:
                        send_mail(f'Certificate Uploaded: {cert.title}', content, settings.DEFAULT_FROM_EMAIL, [t.email], fail_silently=True)
                except Exception:
                    pass
            return redirect('dashboard')
    else:
        form = CertificateForm()
    return render(request, 'upload_certificate.html', {'form': form})

@login_required
def verify_certificate(request, pk, action):
    if not is_approved_teacher(request.user):
        return redirect('dashboard')
    cert = get_object_or_404(Certificate, pk=pk)
    if action == 'approve':
        cert.verified = True
        cert.verified_by = request.user
        cert.feedback = ''
        cert.save()
        Notification.objects.create(user=cert.student, content=f'Your certificate \"{cert.title}\" was approved.')
    else:
        feedback = request.POST.get('feedback','Rejected by teacher')
        cert.verified = False
        cert.feedback = feedback
        cert.verified_by = request.user
        cert.save()
        Notification.objects.create(user=cert.student, content=f'Your certificate \"{cert.title}\" was rejected: {feedback}')
    return redirect('dashboard')


def news_list(request):
    # Publicly visible list of news articles, newest first
    items = News.objects.all().order_by('-created_at')
    return render(request, 'news/news_list.html', {'news_list': items})


def news_detail(request, id):
    n = get_object_or_404(News, pk=id)
    return render(request, 'news/news_detail.html', {'news': n})


@login_required
def add_news(request):
    # Only logged-in users can add news (students, teachers, admins)
    if request.method == 'POST':
        form = NewsForm(request.POST, request.FILES)
        if form.is_valid():
            news = form.save(commit=False)
            news.author = request.user
            # author_role will be set in model.save()
            news.save()
            messages.success(request, 'News posted successfully.')
            return redirect('core:news_list')
    else:
        form = NewsForm()
    return render(request, 'news/add_news.html', {'form': form})


@login_required
def edit_news(request, id):
    n = get_object_or_404(News, pk=id)
    # Only author or staff may edit
    if not (request.user.is_staff or n.author_id == request.user.id):
        return redirect('core:news_detail', id=n.pk)

    if request.method == 'POST':
        form = NewsForm(request.POST, request.FILES, instance=n)
        if form.is_valid():
            news = form.save(commit=False)
            # keep original author
            news.author = n.author
            news.save()
            messages.success(request, 'News updated successfully.')
            return redirect('core:news_detail', id=news.pk)
    else:
        form = NewsForm(instance=n)
    return render(request, 'news/edit_news.html', {'form': form, 'news': n})


@login_required
def delete_news(request, id):
    n = get_object_or_404(News, pk=id)
    # Only author or staff may delete
    if not (request.user.is_staff or n.author_id == request.user.id):
        return redirect('core:news_detail', id=n.pk)

    if request.method == 'POST':
        n.delete()
        messages.success(request, 'News deleted.')
        return redirect('core:news_list')

    return render(request, 'news/confirm_delete.html', {'news': n})


@login_required
def bulk_upload_marks(request):
    # Only approved teachers may use this
    if not is_approved_teacher(request.user):
        return redirect('dashboard')

    results = []
    if request.method == 'POST' and request.FILES.get('file'):
        f = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(filename=BytesIO(f.read()), data_only=True)
            ws = wb.active
            # Expect header row: Student ID | Subject | Marks
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if idx == 1:
                    # skip header
                    continue
                if not row or all([c is None for c in row]):
                    continue
                sid = str(row[0]).strip() if row[0] is not None else ''
                subject = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                marks_val = row[2] if len(row) > 2 else None

                entry = {'student_id': sid, 'exists': False, 'message': ''}

                # Try matching by username or email, then by numeric pk
                student = None
                if sid:
                    student = User.objects.filter(role='student').filter(Q(username=sid) | Q(email=sid)).first()
                    if not student:
                        if sid.isdigit():
                            try:
                                student = User.objects.get(pk=int(sid), role='student')
                            except Exception:
                                student = None

                if not student:
                    entry['message'] = 'Student not found'
                    results.append(entry)
                    continue

                entry['exists'] = True

                # Parse marks
                try:
                    marks_float = float(marks_val) if marks_val is not None and str(marks_val).strip() != '' else None
                except Exception:
                    marks_float = None

                if subject and marks_float is not None:
                    try:
                        Marks.objects.create(student=student, subject=subject, marks_obtained=marks_float, total_marks=100)
                        entry['message'] = 'Imported'
                    except Exception as e:
                        entry['message'] = f'Error saving mark: {str(e)}'
                else:
                    entry['message'] = 'Missing subject or marks'

                results.append(entry)
        except Exception as e:
            results = [{'student_id': '', 'exists': False, 'message': f'Failed to parse file: {str(e)}'}]

    return render(request, 'teachers/bulk_marks_upload.html', {'results': results})


@login_required
def student_availability(request):
    if not is_approved_teacher(request.user):
        return redirect('dashboard')

    results = []
    if request.method == 'POST' and request.FILES.get('file'):
        f = request.FILES['file']
        try:
            wb = openpyxl.load_workbook(filename=BytesIO(f.read()), data_only=True)
            ws = wb.active
            for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if idx == 1:
                    continue
                if not row or all([c is None for c in row]):
                    continue
                sid = str(row[0]).strip() if row[0] is not None else ''
                entry = {'student_id': sid, 'name': '', 'exists': False}
                if sid:
                    student = User.objects.filter(role='student').filter(Q(username=sid) | Q(email=sid)).first()
                    if not student and sid.isdigit():
                        try:
                            student = User.objects.get(pk=int(sid), role='student')
                        except Exception:
                            student = None
                    if student:
                        entry['exists'] = True
                        entry['name'] = student.get_full_name() or student.username
                    else:
                        entry['name'] = ''
                results.append(entry)
        except Exception as e:
            results = [{'student_id': '', 'name': '', 'exists': False, 'error': f'Failed to parse file: {str(e)}'}]

    return render(request, 'teachers/student_availability.html', {'results': results})

@login_required
def add_marks(request):
    if not is_approved_teacher(request.user):
        return redirect('dashboard')
    if request.method == 'POST':
        form = MarksForm(request.POST)
        # restrict selection server-side as well: ensure chosen student belongs to teacher's department
        if form.is_valid():
            student = form.cleaned_data.get('student')
            if student and student.department != request.user.department and not request.user.is_staff:
                messages.error(request, 'You may only add marks for students in your department.')
                return redirect('core:add_marks')
            saved = form.save()
            Notification.objects.create(user=saved.student, content=f'New marks added for {saved.subject}')
            # send email to the student if possible
            try:
                if getattr(settings, 'EMAIL_HOST', None) and saved.student.email:
                    send_mail(f'New Marks: {saved.subject}', f'New marks were added for {saved.subject}. Check your profile for details.', settings.DEFAULT_FROM_EMAIL, [saved.student.email], fail_silently=True)
            except Exception:
                pass
            return redirect('dashboard')
    else:
        form = MarksForm()
        # Restrict selectable students to those in the teacher's department
        form.fields['student'].queryset = User.objects.filter(role='student', is_active=True, department=request.user.department).order_by('first_name', 'last_name', 'email')
    return render(request, 'teachers/add_marks.html', {'form': form})

@login_required
def create_event(request):
    if not is_approved_teacher(request.user):
        return redirect('dashboard')
    if request.method == 'POST':
        form = EventForm(request.POST)
        if form.is_valid():
            ev = form.save(commit=False)
            ev.created_by = request.user
            ev.save()
            # Notify relevant users about the new event
            if ev.scope == 'college':
                targets = User.objects.filter(role='student')
            else:
                targets = User.objects.filter(role='student', department=ev.department)
            content = f'New event posted: "{ev.title}" on {ev.date_from.strftime("%b %d %Y %H:%M")}'
            # create Notification objects and attempt to send email
            for u in targets:
                Notification.objects.create(user=u, content=content)
                # send email if user has email and EMAIL settings present
                try:
                    if getattr(settings, 'EMAIL_HOST', None) and u.email:
                        send_mail(f'New Event: {ev.title}', content, settings.DEFAULT_FROM_EMAIL, [u.email], fail_silently=True)
                except Exception:
                    pass
            return redirect('dashboard')
    else:
        form = EventForm()
    return render(request, 'events/create_event.html', {'form': form})


@login_required
def events_list(request):
    """List events with edit/delete actions for teachers."""
    if not is_approved_teacher(request.user):
        return redirect('dashboard')
    qs = Event.objects.order_by('-date_from')
    return render(request, 'teachers/events_list.html', {'events': qs})


@login_required
def edit_event(request, pk):
    if not is_approved_teacher(request.user):
        return redirect('dashboard')
    ev = get_object_or_404(Event, pk=pk)
    if request.method == 'POST':
        form = EventForm(request.POST, instance=ev)
        if form.is_valid():
            form.save()
            # notify students about important changes
            targets = User.objects.filter(role='student') if ev.scope == 'college' else User.objects.filter(role='student', department=ev.department)
            content = f'Event updated: "{ev.title}" on {ev.date_from.strftime("%b %d %Y %H:%M")}'
            for u in targets:
                Notification.objects.create(user=u, content=content)
                try:
                    if getattr(settings, 'EMAIL_HOST', None) and u.email:
                        send_mail(f'Updated Event: {ev.title}', content, settings.DEFAULT_FROM_EMAIL, [u.email], fail_silently=True)
                except Exception:
                    pass
            return redirect('core:events_list')
    else:
        form = EventForm(instance=ev)
    return render(request, 'teachers/edit_event.html', {'form': form, 'event': ev})


@login_required
def delete_event(request, pk):
    if not is_approved_teacher(request.user):
        return redirect('dashboard')
    ev = get_object_or_404(Event, pk=pk)
    if request.method == 'POST':
        title = ev.title
        ev.delete()
        # notify students that the event was removed
        targets = User.objects.filter(role='student') if ev.scope == 'college' else User.objects.filter(role='student', department=ev.department)
        for u in targets:
            Notification.objects.create(user=u, content=f'Event removed: "{title}"')
            try:
                if getattr(settings, 'EMAIL_HOST', None) and u.email:
                    send_mail(f'Event Cancelled: {title}', f'The event "{title}" has been cancelled.', settings.DEFAULT_FROM_EMAIL, [u.email], fail_silently=True)
            except Exception:
                pass
        return redirect('core:events_list')
    return render(request, 'teachers/confirm_delete_event.html', {'event': ev})


@login_required
def marks_list(request):
    """List all marks for teachers with edit/delete actions."""
    if not is_approved_teacher(request.user):
        return redirect('dashboard')
    # Only show marks for students in this teacher's department (unless staff)
    if request.user.is_staff:
        qs = Marks.objects.select_related('student').order_by('-created_at')
    else:
        qs = Marks.objects.select_related('student').filter(student__department=request.user.department).order_by('-created_at')
    return render(request, 'teachers/marks_list.html', {'marks': qs})


@login_required
def edit_mark(request, pk):
    if not is_approved_teacher(request.user):
        return redirect('dashboard')
    mark = get_object_or_404(Marks, pk=pk)
    # Only allow teachers to edit marks for students in their department
    if not request.user.is_staff and getattr(mark.student, 'department', None) != request.user.department:
        messages.error(request, 'You do not have permission to edit this record.')
        return redirect('dashboard')
    if request.method == 'POST':
        form = MarksForm(request.POST, instance=mark)
        if form.is_valid():
            saved = form.save()
            Notification.objects.create(user=saved.student, content=f'Marks updated for {saved.subject}')
            try:
                if getattr(settings, 'EMAIL_HOST', None) and saved.student.email:
                    send_mail(f'Marks Updated: {saved.subject}', f'Your marks for {saved.subject} were updated.', settings.DEFAULT_FROM_EMAIL, [saved.student.email], fail_silently=True)
            except Exception:
                pass
            return redirect('core:marks_list')
    else:
        form = MarksForm(instance=mark)
        # Ensure the student's own record is selectable even if inactive, and
        # restrict selectable students to the teacher's department.
        dept_qs = User.objects.filter(role='student', is_active=True, department=request.user.department)
        # Allow the specific student to appear even if they are inactive or outside the department
        form.fields['student'].queryset = User.objects.filter(Q(pk=mark.student.pk) | Q(pk__in=dept_qs.values_list('pk', flat=True))).order_by('first_name', 'last_name', 'email')
    return render(request, 'teachers/edit_mark.html', {'form': form, 'mark': mark})


@login_required
def delete_mark(request, pk):
    if not is_approved_teacher(request.user):
        return redirect('dashboard')
    mark = get_object_or_404(Marks, pk=pk)
    # Ensure teacher can only delete marks for students in their department
    if not request.user.is_staff and getattr(mark.student, 'department', None) != request.user.department:
        messages.error(request, 'You do not have permission to delete this record.')
        return redirect('dashboard')
    if request.method == 'POST':
        student = mark.student
        subject = mark.subject
        mark.delete()
        Notification.objects.create(user=student, content=f'Marks removed for {subject}')
        try:
            if getattr(settings, 'EMAIL_HOST', None) and student.email:
                send_mail(f'Marks Deleted: {subject}', f'Your marks for {subject} were deleted by a teacher.', settings.DEFAULT_FROM_EMAIL, [student.email], fail_silently=True)
        except Exception:
            pass
        return redirect('core:marks_list')
    return render(request, 'teachers/confirm_delete_mark.html', {'mark': mark})

@login_required
def notifications(request):
    notes = Notification.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'notifications.html', {'notes': notes})


@login_required
def clear_read_notifications(request):
    """Delete all notifications for the current user that are marked read.

    This endpoint expects a POST request. After deleting, redirects back to
    the notifications page.
    """
    if request.method != 'POST':
        return redirect('core:notifications')
    Notification.objects.filter(user=request.user, read=True).delete()
    messages.success(request, 'Cleared all read notifications.')
    return redirect('core:notifications')

@login_required
def view_profile(request, pk):
    profile_user = get_object_or_404(User, pk=pk)

    # Basic datasets
    marks_qs = Marks.objects.filter(student=profile_user).order_by('created_at')
    certs_qs = Certificate.objects.filter(student=profile_user, verified=True).order_by('uploaded_at')

    # 1) GPA per semester (derive semester from month: Jan-Jun -> S1, Jul-Dec -> S2)
    sem_buckets = OrderedDict()
    for m in marks_qs:
        dt = m.created_at
        year = dt.year
        sem = 1 if dt.month <= 6 else 2
        key = f"{year} S{sem}"
        sem_buckets.setdefault(key, []).append(m.percentage())

    gpa_labels = []
    gpa_data = []
    for key, vals in sem_buckets.items():
        gpa_labels.append(key)
        try:
            gpa_data.append(sum(vals) / len(vals))
        except Exception:
            gpa_data.append(0.0)

    # 2) Certificates over time (monthly counts for last 12 months)
    now = timezone.now()
    cert_labels = []
    cert_counts = []
    # build last 12 months labels (oldest -> newest)
    months = []
    for i in range(11, -1, -1):
        dt = now - timezone.timedelta(days=30 * i)
        months.append((dt.year, dt.month))
    def month_label(y, m):
        return timezone.datetime(y, m, 1).strftime('%b %Y')
    cumulative = 0
    certs_by_month = {(c.uploaded_at.year, c.uploaded_at.month): 0 for c in certs_qs}
    for c in certs_qs:
        key = (c.uploaded_at.year, c.uploaded_at.month)
        certs_by_month[key] = certs_by_month.get(key, 0) + 1

    for (y, m) in months:
        cert_labels.append(month_label(y, m))
        cnt = certs_by_month.get((y, m), 0)
        cert_counts.append(cnt)
        cumulative += cnt

    # NOTE: skill progress proxy removed per request — no skill timeline computed here

    # 4) Leaderboard position history per semester (position among department students for each semester)
    leaderboard_labels = gpa_labels[:]
    leaderboard_positions = []
    if profile_user.department:
        dept_students = User.objects.filter(role='student', department=profile_user.department)
        for key in sem_buckets.keys():
            # parse key like '2023 S1'
            parts = key.split()
            if len(parts) != 2:
                leaderboard_positions.append(None)
                continue
            year = int(parts[0])
            semnum = int(parts[1].lstrip('S'))
            # month range
            if semnum == 1:
                start_month, end_month = 1, 6
            else:
                start_month, end_month = 7, 12
            # compute avg for each student in dept during that semester
            ranks = []
            for s in dept_students:
                s_marks = Marks.objects.filter(student=s, created_at__year=year, created_at__month__gte=start_month, created_at__month__lte=end_month)
                vals = [m.percentage() for m in s_marks]
                avg = (sum(vals) / len(vals)) if vals else None
                ranks.append((s, avg))
            # sort by avg desc, treat None as -inf
            ranks_sorted = sorted(ranks, key=lambda x: (x[1] is not None, x[1] or -1), reverse=True)
            pos = None
            for idx, (s, avg) in enumerate(ranks_sorted, start=1):
                if s.pk == profile_user.pk:
                    pos = idx
                    break
            leaderboard_positions.append(pos or None)
    else:
        leaderboard_positions = [None] * len(leaderboard_labels)

    analytics = {
        'gpa_labels': gpa_labels,
        'gpa_data': [round(x, 2) for x in gpa_data],
        'cert_labels': cert_labels,
        'cert_data': cert_counts,
        'leaderboard_labels': leaderboard_labels,
        'leaderboard_positions': leaderboard_positions,
    }

    context = {
        'profile_user': profile_user,
        'marks': marks_qs,
        'certs': certs_qs,
        'analytics': analytics,
    }
    return render(request, 'profile.html', context)

@login_required
def college_activity(request):
    events = Event.objects.order_by('-date_from')
    certs = Certificate.objects.filter(verified=True).order_by('-uploaded_at')
    marks = Marks.objects.order_by('-created_at')
    return render(request, 'college_activity.html', {'events': events, 'certs': certs, 'marks': marks})


@login_required
def edit_profile(request):
    user: User = request.user
    # Choose the appropriate profile form per role
    if user.role == 'student':
        profile_instance = getattr(user, 'student_profile', None)
        profile_form_class = StudentProfileForm
        profile_context_key = 'student_form'
    else:
        profile_instance = getattr(user, 'teacher_profile', None)
        profile_form_class = TeacherProfileForm
        profile_context_key = 'teacher_form'

    # Ensure profile exists (signals should create, but be defensive)
    if profile_instance is None:
        if user.role == 'student':
            from .models import StudentProfile
            profile_instance = StudentProfile.objects.create(user=user)
        else:
            from .models import TeacherProfile
            profile_instance = TeacherProfile.objects.create(user=user)

    if request.method == 'POST':
        uform = UserEditForm(request.POST, instance=user)
        pform = profile_form_class(request.POST, request.FILES, instance=profile_instance)
        if uform.is_valid() and pform.is_valid():
            uform.save()
            profile_obj = pform.save(commit=False)
            # For StudentProfile, persist computed skills list from clean()
            if hasattr(pform, 'cleaned_data') and 'skills' in pform.cleaned_data:
                profile_obj.skills = pform.cleaned_data['skills']
            profile_obj.save()
            return redirect('core:view_profile', pk=user.pk)
    else:
        uform = UserEditForm(instance=user)
        pform = profile_form_class(instance=profile_instance)

    context = {
        'user_form': uform,
        profile_context_key: pform,
    }
    return render(request, 'profile_edit.html', context)


def check_username(request):
    """AJAX endpoint to check whether a username is available.

    Expects GET param 'username'. Returns JSON: {available: bool, message: str}
    """
    username = request.GET.get('username', '').strip()
    if not username:
        return JsonResponse({'available': False, 'message': 'Enter a username'})

    # Basic validation: length and allowed characters
    if len(username) < 3:
        return JsonResponse({'available': False, 'message': 'Too short (min 3 chars)'})

    if not re.match(r'^[A-Za-z0-9_.-]+$', username):
        return JsonResponse({'available': False, 'message': 'Only letters, numbers, dot, underscore and dash allowed'})

    exists = User.objects.filter(username__iexact=username).exists()
    if exists:
        return JsonResponse({'available': False, 'message': 'Username already taken'})
    return JsonResponse({'available': True, 'message': 'Username is available'})


def check_email(request):
    """AJAX endpoint to validate an email address.

    Checks:
      - non-empty
      - valid email format (using Django's validator)
      - not already used by another user (case-insensitive)

    GET param: 'email'
    Response: { valid: bool, message: str }
    """
    email = request.GET.get('email', '').strip()
    if not email:
        return JsonResponse({'valid': False, 'message': 'Enter an email address'})

    # validate format
    try:
        validate_email(email)
    except ValidationError:
        return JsonResponse({'valid': False, 'message': 'Invalid email format'})

    # allow user's own email when editing profile
    if request.user.is_authenticated and getattr(request.user, 'email', '').lower() == email.lower():
        return JsonResponse({'valid': True, 'message': 'This is your current email'})

    exists = User.objects.filter(email__iexact=email).exists()
    if exists:
        return JsonResponse({'valid': False, 'message': 'Email already registered'})

    return JsonResponse({'valid': True, 'message': 'Email looks good'})


@login_required
def unread_notifications_json(request):
    """Return unread notifications (or notifications after a given id).

    GET params:
      - last_id (optional): only return notifications with id > last_id
    Response: { notifications: [{id, content, created_at}], unread_count: int }
    """
    last_id = request.GET.get('last_id')
    qs = Notification.objects.filter(user=request.user).order_by('created_at')
    if last_id:
        try:
            last_id = int(last_id)
            qs = qs.filter(id__gt=last_id)
        except Exception:
            pass
    # only unread notifications by default
    qs = qs.filter(read=False)
    data = []
    for n in qs:
        data.append({'id': n.id, 'content': n.content, 'created_at': n.created_at.isoformat()})
    unread_count = Notification.objects.filter(user=request.user, read=False).count()
    return JsonResponse({'notifications': data, 'unread_count': unread_count})


@login_required
def mark_notification_read(request):
    """Mark a notification (or all) as read. Expects POST with 'id' or 'all'=1."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST required'}, status=400)
    nid = request.POST.get('id')
    if nid == 'all' or request.POST.get('all') == '1':
        Notification.objects.filter(user=request.user, read=False).update(read=True)
        return JsonResponse({'status': 'ok', 'marked': 'all'})
    try:
        nid = int(nid)
        Notification.objects.filter(user=request.user, id=nid).update(read=True)
        return JsonResponse({'status': 'ok', 'marked': nid})
    except Exception:
        return JsonResponse({'status': 'error', 'message': 'invalid id'}, status=400)


def heartbeat(request):
    """Heartbeat endpoint that returns a stable reload token.

    The endpoint prefers to return the current Git HEAD commit hash when a
    `.git` directory exists. That makes it safe to poll: the token only
    changes when the repository updates (e.g. during development). If Git is
    unavailable, the endpoint falls back to a server timestamp string.

    Clients should poll this endpoint and reload the page when the
    `reload_token` value changes.
    """
    try:
        # Only enable auto-reload token in DEBUG/development to avoid reloads
        # in production. When DEBUG is False return an empty token which the
        # client treats as "no reload".
        from django.conf import settings as _settings
        if not getattr(_settings, 'DEBUG', False):
            return JsonResponse({'reload_token': ''})

        # Try to read the git HEAD commit for a stable development token
        import os
        base = getattr(_settings, 'BASE_DIR', None)
        token = None
        if base:
            git_head = os.path.join(base, '.git', 'HEAD')
            if os.path.exists(git_head):
                try:
                    with open(git_head, 'r', encoding='utf-8') as fh:
                        head = fh.read().strip()
                    if head.startswith('ref:'):
                        ref = head.split(':', 1)[1].strip()
                        ref_path = os.path.join(base, '.git', ref)
                        if os.path.exists(ref_path):
                            with open(ref_path, 'r', encoding='utf-8') as rf:
                                token = rf.read().strip()
                    else:
                        token = head
                except Exception:
                    token = None
        # Fallback in development: use an ISO timestamp so the endpoint still
        # returns a changing token when Git isn't available (e.g., simple
        # deployments).
        if not token:
            token = timezone.now().isoformat()
        return JsonResponse({'reload_token': token})
    except Exception:
        return JsonResponse({'reload_token': ''})


def student_insights(request, pk):
    """Teacher-only view: aggregated insights about a student."""
    if not is_approved_teacher(request.user):
        return redirect('dashboard')

    User = get_user_model()
    student = get_object_or_404(User, pk=pk)
    if student.role != 'student':
        return redirect('dashboard')

    # Marks: overall average and per-subject averages
    marks_qs = Marks.objects.filter(student=student)
    marks_list = list(marks_qs.order_by('-created_at'))
    overall_avg = 0.0
    if marks_list:
        overall_avg = sum(m.percentage() for m in marks_list) / len(marks_list)

    # per-subject averages
    subject_avgs = marks_qs.values('subject').annotate(avg_marks=Avg((F('marks_obtained')/F('total_marks'))*100.0))
    # Serialize subject averages for safe JSON embedding in templates (avoids template tags inside JS)
    subject_avgs_json = []
    for s in subject_avgs:
        try:
            avg = float(s.get('avg_marks') or 0.0)
        except Exception:
            avg = 0.0
        subject_avgs_json.append({'subject': s.get('subject'), 'avg_marks': avg})
    # Certificates
    total_certs = Certificate.objects.filter(student=student).count()
    verified_certs = Certificate.objects.filter(student=student, verified=True).count()

    # Recent posts and comments
    recent_posts = Post.objects.filter(author=student).order_by('-created_at')[:6]
    recent_comments = Comment.objects.filter(author=student).order_by('-created_at')[:6]

    profile = getattr(student, 'student_profile', None)

    context = {
        'student': student,
        'profile': profile,
        'overall_avg': overall_avg,
        'subject_avgs': subject_avgs,
        'subject_avgs_json': subject_avgs_json,
        'total_certs': total_certs,
        'verified_certs': verified_certs,
        'recent_posts': recent_posts,
        'recent_comments': recent_comments,
    }
    return render(request, 'teachers/student_insights.html', context)


@login_required
def toggle_student_active(request, pk):
    """Toggle a student's active state (soft-delete). Teachers can mark students inactive so records are preserved.

    Note: this does NOT hard-delete the User; it flips the built-in `is_active` flag.
    """
    # Limit to teachers (or staff). Adjust if you prefer admins only.
    if not is_approved_teacher(request.user):
        return redirect('dashboard')

    student = get_object_or_404(User, pk=pk, role='student')
    # toggle
    student.is_active = not student.is_active
    student.save()
    if student.is_active:
        messages.success(request, f"Student {student.get_full_name() or student.username} reactivated.")
    else:
        messages.warning(request, f"Student {student.get_full_name() or student.username} deactivated (soft-delete).")
    # Redirect back to teacher dashboard
    return redirect('dashboard')


def event_registrations(request, pk):
    """If an event has a registration_link redirect to it, otherwise show a simple page explaining no link is available.

    This prevents 404s when users visit /core/events/<pk>/registrations/ and provides a place to implement registrations later.
    """
    ev = get_object_or_404(Event, pk=pk)
    # Only redirect to the provided registration URL if registration is still open
    if ev.registration_open:
        # ensure we redirect to an absolute URL
        return redirect(ev.registration_link)
    # If registration is closed or no link provided, render a page explaining the state
    return render(request, 'events/registrations.html', {'event': ev})
